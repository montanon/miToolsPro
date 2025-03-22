import json
import sys
from os import PathLike
from pathlib import Path
from typing import (
    Any,
    Callable,
    Dict,
    Generator,
    Iterable,
    List,
    Optional,
    Sequence,
    Tuple,
    Type,
)

import numpy as np
import openpyxl
from numpy import ndarray
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame, Series

from mitoolspro.exceptions import ArgumentValueError


def iterable_chunks(
    iterable: Iterable, chunk_size: int
) -> Generator[Iterable, None, None]:
    if not isinstance(iterable, (str, list, tuple, bytes)):
        raise TypeError(
            f"Provided iterable of type {type(iterable).__name__} doesn't support slicing."
        )
    for i in range(0, len(iterable), chunk_size):
        yield iterable[i : i + chunk_size]


def dict_from_kwargs(**kwargs: Dict[str, Any]) -> Dict:
    return kwargs


def add_significance(row: Series) -> Series:
    p_value = float(row.split(" ")[1].replace("(", "").replace(")", ""))
    if p_value < 0.001:
        return row + "***"
    elif p_value < 0.01:
        return row + "**"
    elif p_value < 0.05:
        return row + "*"
    else:
        return row


def remove_dataframe_duplicates(dfs: List[DataFrame]) -> List[DataFrame]:
    unique_dfs = []
    for i in range(len(dfs)):
        if not any(dfs[i].equals(dfs[j]) for j in range(i + 1, len(dfs))):
            unique_dfs.append(dfs[i])
    return unique_dfs


def can_convert_to(items: Iterable, type: Type) -> bool:
    try:
        return all(isinstance(type(item), type) for item in items)
    except ValueError:
        return False


def invert_dict(dictionary: Dict) -> Dict:
    return {value: key for key, value in dictionary.items()}


def check_symmetrical_matrix(
    a: ndarray, rtol: Optional[float] = 1e-05, atol: Optional[float] = 1e-08
) -> bool:
    return np.allclose(a, a.T, rtol=rtol, atol=atol)


def unpack_list_of_lists(list_of_lists: List[List]) -> List:
    return [item for sub_list in list_of_lists for item in sub_list]


def auto_adjust_excel_columns_width(excel_path: Path) -> None:
    book = openpyxl.load_workbook(excel_path)
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        auto_adjust_sheet_columns_width(sheet)
    book.save(excel_path)


def auto_adjust_sheet_columns_width(sheet: Worksheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column if cell.value]  # Filter out None values
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = max_length + 1  # Adding a little extra width
        sheet.column_dimensions[
            openpyxl.utils.get_column_letter(column[0].column)
        ].width = adjusted_width


def pretty_dict_str(dictionary: Dict) -> str:
    return json.dumps(dictionary, indent=4, sort_keys=True)


def display_env_variables(
    env_vars: List[Tuple[str, Any]], threshold_mb: float
) -> DataFrame:
    large_vars = []
    for name, value in env_vars:
        size_mb = sys.getsizeof(value) / (1024**2)
        if size_mb > threshold_mb:
            info = f"Type: {type(value).__name__}, ID: {id(value)}"
            if hasattr(value, "__doc__"):
                doc = str(value.__doc__).split("\n")[0]
                info += f", Doc: {doc[:50]}..."
            large_vars.append((name, size_mb, info))
    df = DataFrame(large_vars, columns=["Variable", "Size (MB)", "Info"])
    df.sort_values(by="Size (MB)", ascending=False, inplace=True)
    return df


def sort_dict_keys(
    input_dict: Dict, key: Callable = None, reverse: bool = False
) -> List:
    try:
        sorted_dict = dict(
            sorted(
                input_dict.items(),
                key=key if key else lambda item: item[0],
                reverse=reverse,
            )
        )
        return sorted_dict
    except Exception as e:
        raise ArgumentValueError(f"An error occured shile sorting the dict: {e}")


def get_file_encoding(file: PathLike, fallback: str = "utf-8") -> str:
    try:
        with open(file, "rb") as f:
            raw_data = f.read()
            result = chardet.detect(raw_data)
        encoding = result.get("encoding")
        confidence = result.get("confidence", 0.0)
        if not encoding or confidence < 0.8:
            return fallback
        if encoding.lower() == "ascii":
            return "utf-8"
        return encoding
    except FileNotFoundError:
        raise FileNotFoundError(f"The file '{file}' was not found.")
    except IOError as e:
        raise IOError(f"An error occurred while reading the file '{file}': {e}")


def all_can_be_ints(items: Sequence) -> bool:
    try:
        return all(int(item) is not None for item in items)
    except (ValueError, TypeError):
        return False
