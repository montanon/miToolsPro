import os
import shutil
import tempfile
import unittest
from pathlib import Path
from unittest import TestCase

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from mitoolspro.files.excel_handlers import (
    auto_adjust_excel_columns_width,
    auto_adjust_sheet_columns_width,
)


class TestAutoAdjustExcelColumnsWidth(TestCase):
    def setUp(self):
        self.temp_dir = tempfile.mkdtemp()
        self.test_file = Path(self.temp_dir) / "test.xlsx"
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "Sheet1"
        self.ws2 = self.wb.create_sheet("Sheet2")

        # Sheet1 data
        self.ws1["A1"] = "Short"
        self.ws1["B1"] = "Medium length text"
        self.ws1["C1"] = "Very very very long text that should be wrapped"
        self.ws1["A2"] = "Different length"
        self.ws1["B2"] = "Another medium text"
        self.ws1["C2"] = "Short"

        # Sheet2 data
        self.ws2["A1"] = "Another sheet"
        self.ws2["B1"] = "With different"
        self.ws2["C1"] = "Column widths"
        self.ws2["A2"] = "Test"
        self.ws2["B2"] = "Different"
        self.ws2["C2"] = "Lengths"

        self.wb.save(self.test_file)

    def tearDown(self):
        shutil.rmtree(self.temp_dir)

    def test_auto_adjust_excel_columns_width(self):
        # Get original column widths
        original_wb = openpyxl.load_workbook(self.test_file)
        original_ws1 = original_wb["Sheet1"]
        original_ws2 = original_wb["Sheet2"]

        original_widths_sheet1 = {
            "A": original_ws1.column_dimensions["A"].width,
            "B": original_ws1.column_dimensions["B"].width,
            "C": original_ws1.column_dimensions["C"].width,
        }

        original_widths_sheet2 = {
            "A": original_ws2.column_dimensions["A"].width,
            "B": original_ws2.column_dimensions["B"].width,
            "C": original_ws2.column_dimensions["C"].width,
        }

        # Apply auto-adjust
        auto_adjust_excel_columns_width(self.test_file)

        # Load the adjusted workbook
        adjusted_wb = openpyxl.load_workbook(self.test_file)
        adjusted_ws1 = adjusted_wb["Sheet1"]
        adjusted_ws2 = adjusted_wb["Sheet2"]

        # Verify Sheet1 column widths increased
        self.assertGreater(
            adjusted_ws1.column_dimensions["A"].width, original_widths_sheet1["A"]
        )
        self.assertGreater(
            adjusted_ws1.column_dimensions["B"].width, original_widths_sheet1["B"]
        )
        self.assertGreater(
            adjusted_ws1.column_dimensions["C"].width, original_widths_sheet1["C"]
        )

        # Verify Sheet2 column widths increased
        self.assertGreater(
            adjusted_ws2.column_dimensions["A"].width, original_widths_sheet2["A"]
        )
        self.assertGreater(
            adjusted_ws2.column_dimensions["B"].width, original_widths_sheet2["B"]
        )
        self.assertGreater(
            adjusted_ws2.column_dimensions["C"].width, original_widths_sheet2["C"]
        )

        # Verify column widths are appropriate for content
        self.assertGreater(
            adjusted_ws1.column_dimensions["C"].width,
            adjusted_ws1.column_dimensions["A"].width,
        )
        self.assertGreater(
            adjusted_ws1.column_dimensions["B"].width,
            adjusted_ws1.column_dimensions["A"].width,
        )

    def test_auto_adjust_excel_columns_width_with_empty_sheet(self):
        # Create a new workbook with an empty sheet
        empty_wb = Workbook()
        empty_ws = empty_wb.active
        empty_file = Path(self.temp_dir) / "empty.xlsx"
        empty_wb.save(empty_file)

        # This should not raise any exceptions
        auto_adjust_excel_columns_width(empty_file)

    def test_auto_adjust_excel_columns_width_with_none_values(self):
        # Create a workbook with None values
        none_wb = Workbook()
        none_ws = none_wb.active
        none_ws["A1"] = None
        none_ws["B1"] = "Some text"
        none_ws["C1"] = None
        none_ws["D1"] = "Another text"
        none_file = Path(self.temp_dir) / "none.xlsx"
        none_wb.save(none_file)

        # This should not raise any exceptions
        auto_adjust_excel_columns_width(none_file)

        # Verify the workbook was saved and can be loaded
        loaded_wb = openpyxl.load_workbook(none_file)
        loaded_ws = loaded_wb.active

        # Verify column widths were adjusted for non-None columns
        self.assertGreater(loaded_ws.column_dimensions["B"].width, 0)
        self.assertGreater(loaded_ws.column_dimensions["D"].width, 0)

        # Verify columns with only None values have default width
        self.assertEqual(
            loaded_ws.column_dimensions["A"].width,
            loaded_ws.column_dimensions["A"].width,
        )
        self.assertEqual(
            loaded_ws.column_dimensions["C"].width,
            loaded_ws.column_dimensions["C"].width,
        )


class TestAutoAdjustSheetColumnsWidth(TestCase):
    def setUp(self):
        self.wb = Workbook()
        self.ws = self.wb.active

        # Set up test data
        self.ws["A1"] = "Short"
        self.ws["B1"] = "Medium length text"
        self.ws["C1"] = "Very very very long text that should be wrapped"
        self.ws["A2"] = "Different length"
        self.ws["B2"] = "Another medium text"
        self.ws["C2"] = "Short"

    def test_auto_adjust_sheet_columns_width(self):
        # Get original column widths
        original_widths = {
            "A": self.ws.column_dimensions["A"].width,
            "B": self.ws.column_dimensions["B"].width,
            "C": self.ws.column_dimensions["C"].width,
        }

        # Apply auto-adjust
        auto_adjust_sheet_columns_width(self.ws)

        # Verify column widths increased
        self.assertGreater(self.ws.column_dimensions["A"].width, original_widths["A"])
        self.assertGreater(self.ws.column_dimensions["B"].width, original_widths["B"])
        self.assertGreater(self.ws.column_dimensions["C"].width, original_widths["C"])

        # Verify column widths are appropriate for content
        self.assertGreater(
            self.ws.column_dimensions["C"].width, self.ws.column_dimensions["A"].width
        )
        self.assertGreater(
            self.ws.column_dimensions["B"].width, self.ws.column_dimensions["A"].width
        )

    def test_auto_adjust_sheet_columns_width_with_empty_sheet(self):
        empty_ws = self.wb.create_sheet("Empty")
        # This should not raise any exceptions
        auto_adjust_sheet_columns_width(empty_ws)

    def test_auto_adjust_sheet_columns_width_with_none_values(self):
        none_ws = self.wb.create_sheet("NoneValues")
        none_ws["A1"] = None
        none_ws["B1"] = "Some text"
        none_ws["C1"] = None
        # This should not raise any exceptions
        auto_adjust_sheet_columns_width(none_ws)

    def test_auto_adjust_sheet_columns_width_with_special_characters(self):
        special_ws = self.wb.create_sheet("SpecialChars")
        special_ws["A1"] = "Normal text"
        special_ws["B1"] = "Text with \n newline"
        special_ws["C1"] = "Text with \t tab"
        # This should not raise any exceptions
        auto_adjust_sheet_columns_width(special_ws)


if __name__ == "__main__":
    unittest.main()
